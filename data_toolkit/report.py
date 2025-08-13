from sqlalchemy import create_engine, text
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import json
import re


class ReportBuilder:
    def __init__(self, query_path, mapping_path, db, query_params=None, convert_at_params=False):
        """
        query_params: dict of params for :named_params (e.g., {"current_year": 2025, "previous_year": 2024})
        convert_at_params: if True, will convert @param to :param so SSMS-style SQL can run via SQLAlchemy
        """
        self.db = db
        self.query_path = query_path
        self.mapping = self._load_mapping(mapping_path)   # JSON map stays the same
        self.query_params = query_params or {}
        self.convert_at_params = convert_at_params

        # load data immediately (keeps your original flow)
        self.data = self._query(self.query_path, self.db, self.query_params, self.convert_at_params)

    def run(self):
        # for now, just show what we pulled + mapping available
        print(self.data.columns)
        print(self.mapping)

    def _load_mapping(self, mapping_path):
        raw = json.loads(Path(mapping_path).read_text(encoding='UTF-8'))
        # Your JSON root key typo ("status repot") is handled by taking first value
        return next(iter(raw.values()))

    def _query(self, query_path, db, params: dict, convert_at_params: bool):
        engine = create_engine(db)
        sql = Path(query_path).read_text(encoding='UTF-8')

        # Optional: allow writing @current_year in the .sql file and run via SQLAlchemy
        # Converts @name -> :name (and removes DECLARE blocks if present for testing)
        if convert_at_params:
            # strip DECLARE blocks (handy if your file includes SSMS test harness)
            sql = re.sub(r"(?is)^\s*DECLARE\b.*?;\s*", "", sql)
            sql = re.sub(r"@([A-Za-z_]\w*)", r":\1", sql)

        with engine.begin() as conn:
            df = pd.read_sql_query(text(sql), conn, params=params)
        return df

    def _iter_cell_field_pairs(self):
        for section, inner in self.mapping.items():
            if isinstance(inner, dict):
                for cell, field in inner.items():
                    yield cell, field

    # -------- Optional helpers for batch reporting (multi-row result) --------
    def iter_reports(self, id_col="acct_num"):
        """
        Yields one dict per report row. Perfect for looping and writing one Excel per account.
        """
        for row in self.data.to_dict("records"):
            # suggest a filename based on id_col if present
            filename = f"{row[id_col]}.xlsx" if id_col in row else "report.xlsx"
            yield row, filename

    def _value_for_field(self, row: dict, field: str):
        """Resolve values for both data fields and mapping metadata."""
        if field in row:
            return row[field]
        # simple metadata examples (adjust as you like)
        if field == "header":
            return "Status Report"
        if field == "timeframe":
            cur = self.query_params.get("current_year")
            prev = self.query_params.get("previous_year")
            return f"{prev} vs {cur}"
        return None

    @staticmethod
    def _apply_format(ws, cell_ref: str, value, field_name: str):
        """Basic formatting: percents and currency."""
        cell = ws[cell_ref]
        cell.value = value

        if value is None:
            return

        # percent columns: treat values like 0.06 => 6%
        if isinstance(value, (int, float)) and "percent" in field_name.lower():
            cell.number_format = "0.0%"

        # currency-like: goals and year totals
        if isinstance(value, (int, float)):
            is_goal = field_name.endswith("_goal")
            is_total = field_name.startswith(("previous_year_", "current_year_"))
            if is_goal or is_total:
                cell.number_format = '"$"#,##0;[Red]"$"#,##0'

    def write_excel_for_row(
        self,
        row: dict,
        template_path: str | Path,
        out_path: str | Path,
        sheet_name: str | None = None,
    ):
        """Write one Excel file using a template and the JSON cell mapping."""
        wb = load_workbook(template_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        # fill cells from mapping
        for cell, field in self._iter_cell_field_pairs():
            val = self._value_for_field(row, field)
            self._apply_format(ws, cell, val, field)

        # suggest a filename
        acct = row.get("acct_num", "report")
        out_path = Path(out_path)
        out_path.mkdir(parents=True, exist_ok=True)
        filepath = out_path / f"{acct}_status_report.xlsx"
        wb.save(filepath)
        wb.close()
        return filepath

    def export_all_reports(
        self,
        template_path: str | Path,
        out_dir: str | Path,
        sheet_name: str | None = None,
        id_col: str = "acct_num",
        limit: int | None = None,
    ):
        """Generate one workbook per row. Returns list of paths."""
        paths = []
        count = 0
        for row in self.data.to_dict("records"):
            paths.append(
                self.write_excel_for_row(
                    row=row,
                    template_path=template_path,
                    out_path=out_dir,
                    sheet_name=sheet_name,
                )
            )
            count += 1
            if limit and count >= limit:
                break
        return paths